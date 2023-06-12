import numpy as np
from openpyxl import load_workbook

DEBUG_WIFI_PING = True
DEBUG_TPUT_RETRY = True
DEBUG_RSSI_RETRY = True

infoHead = np.array(['sw version', 'hw version', '2.4/5G BDF', '6G BDF',
                    '5G min t-put',  '5G min rssi',
                    '6G min t-put',  '6G min rssi' ,
                    '24G min t-put', '24G min rssi',
                    'BT min rssi'])
                    
dataHead = np.array(['SN', 'result', 
                    '5G t-put' , '5G rx rate' , 'J16 rssi', 'J17 rssi', 'J18 rssi', 'J19 rssi',
                    '6G t-put' , '6G rx rate' , 'J24 rssi', 'J25 rssi', 'J26 rssi', 'J27 rssi',
                    '24G t-put', '24G rx rate', 'J20 rssi', 'J21 rssi', 'J22 rssi', 'J23 rssi',
                    'BT rssi', 'total test time'])
                    
debugHead = np.array(['5G ping times', '5G tput retry', '5G rssi retry',
                      '6G ping times', '6G tput retry', '6G rssi retry',
                      '24G ping times', '24G tput retry', '24G rssi retry'])

class S4TestCriteria:
    def __init__(self):
        self.sw = ''
        self.hw = ''
        self.BDF_245G = ''
        self.BDF_6G = ''
        self.tput_5G = ''
        self.tput_6G = ''
        self.tput_24G = ''
        self.rssi_5G = ''
        self.rssi_6G = ''
        self.rssi_24G = ''
        self.rssi_BT = ''

class S4TestDetail:
    def __init__(self):
        self.sn = ''
        self.tput_24G = ''
        self.tput_5G = ''
        self.tput_6G = ''
        self.rxRate_24G = ''
        self.rxRate_5G = ''
        self.rxRate_6G = ''
        self.rssi_24G = ['', '', '', '']
        self.rssi_5G = ['', '', '', '']
        self.rssi_6G = ['', '', '', '']
        self.rssi_BT = ''
        self.totalTime = 0
        self.result = ''
        
        self.sumNum = 0

class S4Debug:
    def __init__(self):
        self.pingNum_24G = 0
        self.pingNum_5G = 0
        self.pingNum_6G = 0
        self.tputRetry_24G = -1
        self.tputRetry_5G = -1
        self.tputRetry_6G = -1
        self.rssiRetry_24G = -1
        self.rssiRetry_5G = -1
        self.rssiRetry_6G = -1
        
        self.sumNum = 0

def get_info(file):
    this = S4TestCriteria()
    with open(file, "r") as f:
        line = f.readline()
        
        while line:
            if line.find('FW Version:') != -1 and not 'WLAN' in line:
                this.sw = line.partition(':')[2].strip()
            elif line.find('HW Version:') != -1:
                this.hw = line.partition(':')[2].strip()
            elif line.find('  /lib/firmware/IPQ8074/bdwlan.b294') != -1:
                this.BDF_245G = line.partition('  ')[0].strip()
            elif line.find('  /lib/firmware/qcn9000/bdwlan.ba4') != -1:
                this.BDF_6G = line.partition('  ')[0].strip()
            elif line.find('WIFI5G_TPUT_MIN') != -1:
                this.tput_5G = line.partition(':')[2].strip()
            elif line.find('WIFI6G_TPUT_MIN') != -1:
                this.tput_6G = line.partition(':')[2].strip()
            elif line.find('WIFI24G_TPUT_MIN') != -1:
                this.tput_24G = line.partition(':')[2].strip()
            elif line.find('WIFI5G_RSSI_MIN') != -1:
                this.rssi_5G = line.partition(':')[2].strip()
            elif line.find('WIFI6G_RSSI_MIN') != -1:
                this.rssi_6G = line.partition(':')[2].strip()
            elif line.find('WIFI24G_RSSI_MIN') != -1:
                this.rssi_24G = line.partition(':')[2].strip()
            elif line.find('BT_RSSI_MIN') != -1:
                this.rssi_BT = line.partition(':')[2].strip()
            line = f.readline()
        f.close()
        
    arr = np.array([this.sw, this.hw, this.BDF_245G, this.BDF_6G,
                    this.tput_5G, this.rssi_5G,
                    this.tput_6G, this.rssi_6G,
                    this.tput_24G, this.rssi_24G,
                    this.rssi_BT])

    del this
    return np.vstack([infoHead, arr])

def do_parsing(file):
    pingTemp = 0
    getRX = False
    wifiRadio = ''
    this = S4TestDetail()
    thisDbg = S4Debug()
    
    with open(file, "r") as f:
        line = f.readline()
        
        while line:
            if line.find('sys_serialno=') != -1:
                this.sn = line.partition('=')[2].strip()
            elif line.find('__get_sta_count') != -1:
                if wifiRadio == '5G':
                    wifiRadio = '6G'
                elif wifiRadio == '6G':
                    wifiRadio = '24G'
                else:
                    wifiRadio = '5G'
            elif line.find('1 packets transmitted') != -1:
                pingTemp += 1
            elif line.find('------------------------------------------------------------') != -1:
                if thisDbg.pingNum_5G == 0:
                    thisDbg.pingNum_5G = pingTemp
                elif thisDbg.pingNum_6G == 0:
                    thisDbg.pingNum_6G = pingTemp
                elif thisDbg.pingNum_24G == 0:
                    thisDbg.pingNum_24G = pingTemp
                pingTemp = 0
            elif line.find('__do_tput') != -1:
                if wifiRadio == '5G':
                    thisDbg.tputRetry_5G += 1
                elif wifiRadio == '6G':
                    thisDbg.tputRetry_6G += 1
                elif wifiRadio == '24G':
                    thisDbg.tputRetry_24G += 1
            elif line.find('Server listening on UDP port 5001') != -1:
                this.sumNum += 1
            elif line.find('[SUM]') != -1:
                if wifiRadio == '24G':
                    this.tput_24G = line.split()[6].strip()
                elif wifiRadio == '5G':
                    this.tput_5G = line.split()[6].strip()
                elif wifiRadio == '6G':
                    this.tput_6G = line.split()[6].strip()
            elif line.find('cat /tmp/wifi_rssi.txt') != -1:
                if wifiRadio == '5G':
                    thisDbg.rssiRetry_5G += 1
                elif wifiRadio == '6G':
                    thisDbg.rssiRetry_6G += 1
                elif wifiRadio == '24G':
                    thisDbg.rssiRetry_24G += 1
            elif line.find('__get_rxRate') != -1:
                getRX = True
            elif line.strip().endswith('M') and line.strip().removesuffix('M').isdigit() and getRX:
                if wifiRadio == '24G':
                    this.rxRate_24G = line.strip()
                    getRX = True
                elif wifiRadio == '5G':
                    this.rxRate_5G = line.strip()
                elif wifiRadio == '6G':
                    this.rxRate_6G = line.strip()
                getRX = False
            elif line.find('J16 max rssi=') != -1:
                this.rssi_5G[0] = line.rpartition('=')[2].strip()
            elif line.find('J17 max rssi=') != -1:
                this.rssi_5G[1] = line.rpartition('=')[2].strip()
            elif line.find('J18 max rssi=') != -1:
                this.rssi_5G[2] = line.rpartition('=')[2].strip()
            elif line.find('J19 max rssi=') != -1:
                this.rssi_5G[3] = line.rpartition('=')[2].strip()
            elif line.find('J24 max rssi=') != -1:
                this.rssi_6G[0] = line.rpartition('=')[2].strip()
            elif line.find('J25 max rssi=') != -1:
                this.rssi_6G[1] = line.rpartition('=')[2].strip()
            elif line.find('J26 max rssi=') != -1:
                this.rssi_6G[2] = line.rpartition('=')[2].strip()
            elif line.find('J27 max rssi=') != -1:
                this.rssi_6G[3] = line.rpartition('=')[2].strip()
            elif line.find('J20 max rssi=') != -1:
                this.rssi_24G[0] = line.rpartition('=')[2].strip()
            elif line.find('J21 max rssi=') != -1:
                this.rssi_24G[1] = line.rpartition('=')[2].strip()
            elif line.find('J22 max rssi=') != -1:
                this.rssi_24G[2] = line.rpartition('=')[2].strip()
            elif line.find('J23 max rssi=') != -1:
                this.rssi_24G[3] = line.rpartition('=')[2].strip()
            elif line.find('min rssi = ') != -1:
                this.rssi_BT = line.rpartition('=')[2].strip()
            elif line.find('Total Test Time:') != -1:
                this.totalTime = line.rpartition(':')[2].strip().removesuffix('s')
            elif line.find('SUMMARY:') != -1:
                this.result = line.partition(':')[2].strip()
            line = f.readline()
        f.close()
    
    # if this.sumNum >= 3:
    #     print(file + ' ' + str(thisDbg.pingNum_5G) + ' ' + str(thisDbg.pingNum_6G) + ' ' + str(thisDbg.pingNum_24G) + ' ' + str(this.sumNum))
    
    arr = np.array([this.sn, this.result, 
                    this.tput_5G,  this.rxRate_5G,  this.rssi_5G[0],  this.rssi_5G[1],  this.rssi_5G[2],  this.rssi_5G[3],
                    this.tput_6G,  this.rxRate_6G,  this.rssi_6G[0],  this.rssi_6G[1],  this.rssi_6G[2],  this.rssi_6G[3],
                    this.tput_24G, this.rxRate_24G, this.rssi_24G[0], this.rssi_24G[1], this.rssi_24G[2], this.rssi_24G[3],
                    this.rssi_BT, this.totalTime])
    
    arrDbg = np.array([thisDbg.pingNum_5G, thisDbg.tputRetry_5G, thisDbg.rssiRetry_5G,
                       thisDbg.pingNum_6G, thisDbg.tputRetry_6G, thisDbg.rssiRetry_6G,
                       thisDbg.pingNum_24G, thisDbg.tputRetry_24G, thisDbg.rssiRetry_24G])
    del this
    return arr, arrDbg
    
def create_report(reportTemplate, reportName, workSheet, headArray, dataArray, debugArray):
    wb = load_workbook(filename = reportTemplate)
    ws = wb[workSheet]
    x = 1
    y = 0
    
    for row in range(10, 10 + dataArray.shape[0] - 1, 1):
        for col in range(1, 1 + dataArray.shape[1], 1):
            if col > 2 and dataArray[x, y].lstrip('-').isdigit():
                ws.cell(row = row, column = col, value = int(dataArray[x, y]))
            elif col > 2 and dataArray[x, y].partition('.')[0].isdigit() and dataArray[x, y].partition('.')[2].isdigit():
                ws.cell(row = row, column = col, value = float(dataArray[x, y]))
            else:
                ws.cell(row = row, column = col, value = dataArray[x, y])
            y += 1
        x += 1
        y = 0

    x = 1
    y = 0
    for row in range(10, 10 + debugArray.shape[0] - 1, 1):
        for col in range(24, 24 + debugArray.shape[1], 1):
            ws.cell(row = row, column = col, value = debugArray[x, y])
            y += 1
        x += 1
        y = 0
        
    wb.save(reportName)